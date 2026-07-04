#!/usr/bin/env python3
"""Genera la matriz Excel de status LEYENDO EL TABLERO (única fuente de verdad).

Uso: python reportes/generar_matriz_status.py
Salida: reportes/Status_Proyectos_Behavioral_Design.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from beholder_tools import parse_tablero

SALIDA = Path(__file__).resolve().parent / "Status_Proyectos_Behavioral_Design.xlsx"

COLUMNAS = [
    ("Clave", 8), ("Épica", 26), ("Iniciativa", 44), ("Estado", 12),
    ("Status del proyecto", 46), ("Intervención conductual", 52),
    ("Fecha de inicio", 13), ("Fecha de cierre", 13), ("Monedas 🪙", 10),
    ("Desglose 🪙", 22), ("Behavioral designers", 22), ("Otros perfiles", 15),
    ("Riesgos", 40), ("Impacto esperado", 44),
]

COLOR_ESTADO = {
    "Done": "C6EFCE", "In Progress": "BDD7EE", "In Review": "FFE699",
    "To Do": "F2F2F2", "Backlog": "E1D5E7",
}


def main():
    quests = parse_tablero()
    wb = Workbook()
    ws = wb.active
    ws.title = "Status BD Q3-2026"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    borde = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for col, (titulo, ancho) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 28

    for fila, q in enumerate(quests, 2):
        desglose = " · ".join(f"{p}: {n:g}" for p, n in q["monedas"].items()) or "—"
        valores = [
            q["clave"], f"{q['epica']['id']} · {q['epica']['nombre']}", q["nombre"],
            q["estado"], q["status"], q["intervencion"],
            q["inicio"].strftime("%d/%m/%Y") if q["inicio"] else "—",
            q["cierre"].strftime("%d/%m/%Y") if q["cierre"] else "—",
            q["monedas_total"] if q["monedas_total"] else "—",
            desglose, q["bds"], q["perfiles"], q["riesgos"], q["impacto"],
        ]
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.border = borde
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        estado_cell = ws.cell(row=fila, column=4)
        if q["estado"] in COLOR_ESTADO:
            estado_cell.fill = PatternFill("solid", fgColor=COLOR_ESTADO[q["estado"]])
            estado_cell.alignment = Alignment(horizontal="center", vertical="center")
        if "🚩" in q["riesgos"]:
            ws.cell(row=fila, column=13).font = Font(color="C00000", bold=True)

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{len(quests) + 1}"
    wb.save(SALIDA)
    print(f"Excel regenerado desde el tablero: {SALIDA.name} ({len(quests)} iniciativas).")


if __name__ == "__main__":
    main()
